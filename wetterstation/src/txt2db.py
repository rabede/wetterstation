#!/usr/bin/env python3
'''
Created on 27.12.2016

@author: bell
'''
import getpass
import os
import sqlite3

import keys
import openpyxl


conns = sqlite3.connect(**keys.SQLITECONFIG)
curs = conns.cursor()
data_dir = "/home/" + getpass.getuser() + "/Dropbox/Apps/51381/"

def save_data(curs, data):
    try:
        curs.execute('''INSERT OR IGNORE into wetter_raw (timestamp, temperatur, humidity, windspeed, downfall, rain)       
                       VALUES (?,?,?,?,?,?)''',     (data[0], data[1], data[2], data[3], data[4], data[5]))
    except:
        print('Fehler: ', data)


#importformat ['2015-02-15 07:12', '-2.3', '95.0', '0.0', 0, '0'] von raspberry:
def get_wetterpi(file):
        text = open(file)
        for line in text:
            cleanline = line.strip("[]\n").replace("'", "")
            data = cleanline.split(",")
            save_data(curs, data)

def get_google(file):
    data = [""]*6
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name('Tabellenblatt1')
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row = row, column = 1).value == None:
            continue
        for col in range(6):
            data[col] = sheet.cell(row = row, column =  col+1).value

        if data[5] == 'ja':
            data[5] = 1
        else:
            data[5] = 0
 
        save_data(curs, data)
      
def reset_filenames(path):
    for file in os.listdir(path):
        os.rename(file, file.replace('.read', ''))

def loop_dir(path):
    ''' Loop over the files in the working directory.'''
    for file in os.listdir(path):
        if file.endswith('.txt'):
            get_wetterpi(file)
        elif file.endswith('.xlsx'):
            get_google(file)
        else:
            continue
        os.rename(file, file + '.read')
        conns.commit()

def start():
    os.chdir(data_dir)
    #reset_filenames(data_dir)
    loop_dir(data_dir)
    
    conns.commit()
    conns.close()

    
if __name__ == '__main__':    
    start()